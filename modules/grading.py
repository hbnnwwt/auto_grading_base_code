import openpyxl

from modules.pipeline import LAYOUT


def _classify_question(q_num):
    """根据题号判断题型，从配置动态获取边界。"""
    _ch = LAYOUT['choice']
    _ju = LAYOUT['judge']
    choice_end = _ch['question_start'] + _ch['question_count'] - 1
    judge_end = _ju['question_start'] + _ju['question_count'] - 1
    if _ch['question_start'] <= q_num <= choice_end:
        return 'choice'
    elif _ju['question_start'] <= q_num <= judge_end:
        return 'judge'
    else:
        return 'essay'


class EssayGraderBase:
    """简答题评分器抽象基类。子类需实现 score() 方法。"""

    def score(self, question, reference, student_answer, max_score):
        raise NotImplementedError


class DefaultEssayGrader(EssayGraderBase):
    """默认实现：简答题返回 0 分，标注需手动评分。"""

    def score(self, question, reference, student_answer, max_score):
        return 0, max_score, "需手动评分"


class GradingService:
    """评分模块：将识别结果与标准答案比对，生成评分报告。"""

    def __init__(self, answer_key, essay_grader=None,
                 choice_score=None, judge_score=None, essay_max_score=None):
        """
        Args:
            answer_key: 标准答案字典
                {'choice': {1:'A', 2:'C', ...},
                 'judge':  {21:'T', 22:'F', ...},
                 'essay':  {31:'参考答案文本'}}
            essay_grader: 简答题评分器，默认使用 DefaultEssayGrader
            choice_score: 每道选择题分值，默认从配置加载
            judge_score: 每道判断题分值，默认从配置加载
            essay_max_score: 简答题满分，默认从配置加载
        """
        self.answer_key = answer_key
        self.essay_grader = essay_grader or DefaultEssayGrader()
        _scoring = LAYOUT.get('scoring', {})
        self.choice_score = choice_score if choice_score is not None else _scoring.get('choice_score', 3)
        self.judge_score = judge_score if judge_score is not None else _scoring.get('judge_score', 2)
        self.essay_max_score = essay_max_score if essay_max_score is not None else _scoring.get('essay_max_score', 20)

    @property
    def max_total(self):
        """动态计算满分。"""
        n_choice = len(self.answer_key.get('choice', {}))
        n_judge = len(self.answer_key.get('judge', {}))
        n_essay = len(self.answer_key.get('essay', {}))
        return (n_choice * self.choice_score
                + n_judge * self.judge_score
                + n_essay * self.essay_max_score)

    @classmethod
    def from_xlsx(cls, path):
        """从 参考答案.xlsx 加载标准答案。

        Args:
            path: xlsx 文件路径

        Returns:
            GradingService 实例
        """
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        answer_key = {'choice': {}, 'judge': {}, 'essay': {}}
        for col in range(2, ws.max_column + 1):
            q_num = ws.cell(row=1, column=col).value
            answer = ws.cell(row=2, column=col).value
            if q_num is None or answer is None:
                continue
            try:
                q_num = int(q_num)
            except (ValueError, TypeError):
                continue
            q_type = _classify_question(q_num)
            answer_key[q_type][q_num] = answer

        _scoring = LAYOUT.get('scoring', {})
        return cls(answer_key,
                   choice_score=_scoring.get('choice_score', 3),
                   judge_score=_scoring.get('judge_score', 2),
                   essay_max_score=_scoring.get('essay_max_score', 20))

    def grade(self, recognized_answers):
        """将识别结果与标准答案比对，计算得分。

        Args:
            recognized_answers: 各题型识别结果
                {'choice': {1:'A', 2:'B', ...},
                 'judge':  {21:'T', 22:'F', ...},
                 'essay':  {31:'识别文本'}}

        Returns:
            dict: 包含各题得分、总分、错误明细的评分报告

        启发式问题：如何处理学生未作答的题目（识别结果为 None）？
        简答题与客观题的评分逻辑有何本质区别？
        """
        raise NotImplementedError("请实现评分逻辑")

    def generate_report(self, result):
        """生成可读的评分报告。

        Args:
            result: grade() 的返回值

        Returns:
            str: 格式化的评分报告文本
        """
        lines = [f"总分: {result['total']}/{self.max_total}", ""]

        n_choice = len(self.answer_key.get('choice', {}))
        n_judge = len(self.answer_key.get('judge', {}))
        choice_correct = sum(1 for d in result['choice'].values() if d['score'] > 0)
        lines.append(f"选择题: {choice_correct}/{n_choice} ({choice_correct * self.choice_score}分)")
        for q in sorted(result['choice']):
            d = result['choice'][q]
            mark = '✅' if d['score'] > 0 else '❌'
            lines.append(f"  第{q}题: {d['given'] or '未答'} (正确:{d['correct']}) {mark}")

        judge_correct = sum(1 for d in result['judge'].values() if d['score'] > 0)
        lines.append(f"判断题: {judge_correct}/{n_judge} ({judge_correct * self.judge_score}分)")
        for q in sorted(result['judge']):
            d = result['judge'][q]
            mark = '✅' if d['score'] > 0 else '❌'
            lines.append(f"  第{q}题: {d['given'] or '未答'} (正确:{d['correct']}) {mark}")

        essay_detail = result.get('essay_detail', {})
        if essay_detail:
            lines.append(f"简答题: {result.get('essay_total', 0)}/{len(self.answer_key.get('essay', {})) * self.essay_max_score}分")
            for q in sorted(essay_detail):
                d = essay_detail[q]
                lines.append(f"  第{q}题: {d['score']}/{d['max_score']}分 — {d['feedback']}")

        return '\n'.join(lines)

    def save_result_xlsx(self, template_path, output_path,
                         student_results):
        """将识别结果按 结果.xlsx 模板格式输出。

        Args:
            template_path: 结果.xlsx 模板路径
            output_path: 输出文件路径
            student_results: 列表，每个元素为 (学号, recognized_answers)

        启发式问题：如何根据题号判断题目类型？
        Excel 公式与直接写入计算结果各有什么优劣？
        """
        raise NotImplementedError("请实现结果导出")


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        service = GradingService.from_xlsx(sys.argv[1])
        print(f"已加载标准答案: 选择题{len(service.answer_key['choice'])}题, "
              f"判断题{len(service.answer_key['judge'])}题, "
              f"简答题{len(service.answer_key['essay'])}题")
