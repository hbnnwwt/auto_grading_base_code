"""评分模块测试。

验证 GradingService 的评分逻辑、报告生成和 xlsx 输出。
运行: pytest tests/test_grading.py -v
"""

import os

import openpyxl
import pytest

from modules.grading import GradingService


# ---------------------------------------------------------------------------
# fixture: 标准答案
# ---------------------------------------------------------------------------

@pytest.fixture
def answer_key():
    """构造标准答案字典。"""
    return {
        'choice': {i: chr(ord('A') + (i % 4)) for i in range(1, 21)},
        'judge': {i: 'T' if i % 2 == 1 else 'F' for i in range(21, 31)},
        'essay': {31: '这是一段参考答案文本'},
    }


@pytest.fixture
def service(answer_key):
    """基于标准答案创建 GradingService。"""
    return GradingService(answer_key)


# ---------------------------------------------------------------------------
# GradingService.grade
# ---------------------------------------------------------------------------

class TestGrade:
    def test_all_correct(self, service):
        """全部答对应得满分。"""
        recognized = {
            'choice': {i: chr(ord('A') + (i % 4)) for i in range(1, 21)},
            'judge': {i: 'T' if i % 2 == 1 else 'F' for i in range(21, 31)},
            'essay': {31: '测试文本'},
        }
        result = service.grade(recognized)
        assert result['total'] == 80  # 20*3 + 10*2

    def test_all_wrong(self, service):
        """全部答错应得 0 分。"""
        recognized = {
            'choice': {i: 'X' for i in range(1, 21)},
            'judge': {i: 'X' for i in range(21, 31)},
            'essay': {},
        }
        result = service.grade(recognized)
        assert result['total'] == 0

    def test_partial_correct(self, service):
        """部分答对的得分。"""
        # 答对第 1 题(答案B)和第 21 题(答案T)
        recognized = {
            'choice': {1: 'B', 2: 'X'},  # 第1题对，第2题错
            'judge': {21: 'T', 22: 'X'},  # 第21题对，第22题错
            'essay': {},
        }
        result = service.grade(recognized)
        assert result['total'] == 5  # 1*3 + 1*2

    def test_empty_answers(self, service):
        """未作答应得 0 分。"""
        result = service.grade({'choice': {}, 'judge': {}, 'essay': {}})
        assert result['total'] == 0

    def test_result_has_details(self, service):
        """结果应包含各题得分明细。"""
        recognized = {
            'choice': {1: 'A'},
            'judge': {21: 'T'},
            'essay': {},
        }
        result = service.grade(recognized)
        assert 'choice' in result
        assert 'judge' in result
        assert 'total' in result


# ---------------------------------------------------------------------------
# GradingService.generate_report
# ---------------------------------------------------------------------------

class TestGenerateReport:
    def test_report_is_string(self, service):
        """报告应为字符串。"""
        recognized = {'choice': {1: 'A'}, 'judge': {}, 'essay': {}}
        result = service.grade(recognized)
        report = service.generate_report(result)
        assert isinstance(report, str)

    def test_report_contains_score(self, service):
        """报告中应包含总分。"""
        recognized = {'choice': {1: 'A'}, 'judge': {21: 'T'}, 'essay': {}}
        result = service.grade(recognized)
        report = service.generate_report(result)
        assert str(result['total']) in report

    def test_report_not_empty(self, service):
        """报告应非空。"""
        recognized = {'choice': {}, 'judge': {}, 'essay': {}}
        result = service.grade(recognized)
        report = service.generate_report(result)
        assert len(report) > 0


# ---------------------------------------------------------------------------
# GradingService.from_xlsx
# ---------------------------------------------------------------------------

class TestFromXlsx:
    def test_load_answer_key(self):
        """从项目自带的 参考答案.xlsx 加载。"""
        xlsx_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), '参考答案.xlsx'
        )
        if not os.path.exists(xlsx_path):
            pytest.skip("参考答案.xlsx 不存在")

        service = GradingService.from_xlsx(xlsx_path)
        assert len(service.answer_key['choice']) == 20
        assert len(service.answer_key['judge']) == 10
        assert len(service.answer_key['essay']) >= 1

    def test_answer_key_values(self):
        """验证加载的答案格式正确。"""
        xlsx_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), '参考答案.xlsx'
        )
        if not os.path.exists(xlsx_path):
            pytest.skip("参考答案.xlsx 不存在")

        service = GradingService.from_xlsx(xlsx_path)
        # 选择题答案应为 A/B/C/D
        for q, ans in service.answer_key['choice'].items():
            assert ans in ['A', 'B', 'C', 'D'], f"第{q}题答案格式错误: {ans}"
        # 判断题答案应为 T/F
        for q, ans in service.answer_key['judge'].items():
            assert ans in ['T', 'F'], f"第{q}题答案格式错误: {ans}"


# ---------------------------------------------------------------------------
# GradingService.save_result_xlsx
# ---------------------------------------------------------------------------

class TestSaveResultXlsx:
    def test_save_creates_file(self, service, tmp_path):
        """save_result_xlsx 应创建输出文件。"""
        template_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), '结果.xlsx'
        )
        if not os.path.exists(template_path):
            pytest.skip("结果.xlsx 模板不存在")

        output_path = str(tmp_path / "output.xlsx")
        student_results = [
            (2025811001, {
                'choice': {1: 'A', 2: 'B'},
                'judge': {21: 'T'},
                'essay': {},
            }),
        ]

        service.save_result_xlsx(template_path, output_path, student_results)
        assert os.path.exists(output_path)

    def test_output_has_student_id(self, service, tmp_path):
        """输出文件中应包含学号。"""
        template_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), '结果.xlsx'
        )
        if not os.path.exists(template_path):
            pytest.skip("结果.xlsx 模板不存在")

        output_path = str(tmp_path / "output.xlsx")
        student_id = 2025811099
        student_results = [
            (student_id, {'choice': {}, 'judge': {}, 'essay': {}}),
        ]

        service.save_result_xlsx(template_path, output_path, student_results)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        ids = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        assert student_id in ids


# ---------------------------------------------------------------------------
# EssayGraderBase / DefaultEssayGrader
# ---------------------------------------------------------------------------

class TestEssayGraderBase:
    def test_default_grader_returns_zero(self):
        """默认评分器应返回 0 分和手动评分标记。"""
        from modules.grading import DefaultEssayGrader
        grader = DefaultEssayGrader()
        score, max_score, feedback = grader.score("题目", "参考答案", "学生答案", 10)
        assert score == 0
        assert max_score == 10
        assert "手动" in feedback

    def test_custom_grader_integration(self, service):
        """自定义 essay_grader 应被 grade() 调用。"""
        from modules.grading import GradingService, EssayGraderBase

        class StubGrader(EssayGraderBase):
            def score(self, question, reference, student_answer, max_score):
                return max_score, max_score, "满分"

        svc = GradingService(service.answer_key, essay_grader=StubGrader())
        recognized = {
            'choice': {1: 'A'},
            'judge': {},
            'essay': {31: "学生答案"},
        }
        result = svc.grade(recognized)
        assert result['essay_detail'][31]['score'] == 20
        assert result['total'] > 3

    def test_grade_without_essay_grader(self, service):
        """无 essay_grader 时简答题得 0 分。"""
        recognized = {
            'choice': {},
            'judge': {},
            'essay': {31: "学生答案"},
        }
        result = service.grade(recognized)
        assert result['essay_detail'][31]['score'] == 0
